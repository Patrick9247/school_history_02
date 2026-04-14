#!/usr/bin/env python3
"""
导入专业数据从 Excel 文件到数据库
"""
import sys
import os

# 添加 backend 目录到路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from database import get_db_connection, init_db

def import_majors_from_excel(excel_path: str):
    """从 Excel 文件导入专业数据"""
    
    # 初始化数据库
    init_db()
    
    # 连接数据库
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 获取年份映射 {year: id}
    cursor.execute("SELECT id, year FROM years")
    year_map = {row['year']: row['id'] for row in cursor.fetchall()}
    print(f"已加载 {len(year_map)} 个年份")
    
    # 获取学院映射 {name: id}
    cursor.execute("SELECT id, name FROM colleges")
    college_map = {row['name']: row['id'] for row in cursor.fetchall()}
    print(f"已加载 {len(college_map)} 个学院")
    
    # 读取 Excel 文件
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # 导入统计
    imported = 0
    skipped = 0
    errors = []
    
    # 遍历数据行 (跳过表头)
    for row_num in range(2, ws.max_row + 1):
        try:
            # 读取各列数据
            major_name = ws.cell(row_num, 1).value  # 专业名称
            major_code = ws.cell(row_num, 2).value  # 专业代码
            year_value = ws.cell(row_num, 3).value  # 年度
            level = ws.cell(row_num, 4).value       # 学制
            school_name = ws.cell(row_num, 5).value  # 学校名称
            department = ws.cell(row_num, 6).value  # 原所在院系
            college_name = ws.cell(row_num, 7).value  # 现归属学院
            
            # 跳过空数据行
            if not major_name:
                skipped += 1
                continue
            
            # 处理年份
            if year_value:
                if isinstance(year_value, str):
                    year = int(year_value)
                elif isinstance(year_value, (int, float)):
                    year = int(year_value)
                else:
                    year = 1956  # 默认值
            else:
                year = 1956
            
            year_id = year_map.get(year)
            if not year_id:
                # 如果年份不存在，跳过
                errors.append(f"行{row_num}: 年份 {year} 不存在")
                skipped += 1
                continue
            
            # 处理学院
            college_id = None
            if college_name:
                college_id = college_map.get(college_name)
                if not college_id:
                    # 如果学院不存在，尝试模糊匹配
                    for name, cid in college_map.items():
                        if college_name in name or name in college_name:
                            college_id = cid
                            break
            
            # 插入数据
            cursor.execute("""
                INSERT INTO majors 
                (year_id, college_id, name, code, category, description, level, department)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                year_id,
                college_id,
                str(major_name).strip(),
                str(major_code).strip() if major_code else None,
                college_name,
                None,  # description
                level,
                str(department).strip() if department else None
            ))
            
            imported += 1
            
            if imported % 500 == 0:
                conn.commit()
                print(f"已导入 {imported} 条...")
                
        except Exception as e:
            errors.append(f"行{row_num}: {str(e)}")
            skipped += 1
    
    # 提交事务
    conn.commit()
    conn.close()
    
    # 输出结果
    print("\n" + "="*50)
    print("导入完成!")
    print(f"成功导入: {imported} 条")
    print(f"跳过: {skipped} 条")
    if errors:
        print(f"错误: {len(errors)} 条")
        for err in errors[:10]:
            print(f"  - {err}")
        if len(errors) > 10:
            print(f"  ... 还有 {len(errors) - 10} 条错误")
    print("="*50)

if __name__ == "__main__":
    excel_path = os.path.join(os.path.dirname(__file__), "..", "data", "import_major.xlsx")
    excel_path = os.path.abspath(excel_path)
    
    if not os.path.exists(excel_path):
        print(f"错误: 文件不存在 {excel_path}")
        sys.exit(1)
    
    print(f"Excel 文件: {excel_path}")
    import_majors_from_excel(excel_path)
